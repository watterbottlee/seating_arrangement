import openpyxl

print("program has started successfully")
# Get branch name, total number of students, and starting roll number
branch_name = input("Enter branch name (e.g. CSE): ")
total_students = int(input("Enter total number of students: "))
starting_roll_no = int(input("Enter starting roll number (e.g. 22027101): "))

try: 
         # Initialize variables
        students_left = total_students
        current_roll_no = starting_roll_no

        # Create an Excel file
        wb = openpyxl.Workbook()
        ws = wb.active

        # Set the title row
        ws['A1'] = "Room"
        ws['B1'] = "Row 1"
        ws['C1'] = "Row 2"
        ws['D1'] = "Row 3"
        ws['E1'] = "Row 4"
        ws['F1'] = "Row 5"

        excel_row = 2
        room_no = 1

        while students_left > 0:
            # Get room details
            room_name = input(f"Enter room number {room_no}'s name (e.g. F21): ")
            benches_available = int(input(f"Enter number of benches available in room {room_name}: "))
            rows_to_make = int(input(f"Enter number of rows to make in room {room_name} (3 or 4): "))

            # Calculate number of benches to fill in this room
            benches_to_fill = min(benches_available, students_left)

            # Calculate benches per row
            benches_per_row = benches_to_fill // rows_to_make
            extra_benches = benches_to_fill % rows_to_make
            
            row_benches = [benches_per_row] * rows_to_make
            for i in range(extra_benches):
                row_benches[i] += 1

            # Fill the room in the Excel sheet
            start_row = excel_row
            for col in range(rows_to_make):
                for row in range(row_benches[col]):
                    ws.cell(row=excel_row, column=1).value = room_name
                    ws.cell(row=excel_row, column=col+2).value = current_roll_no
                    current_roll_no += 1
                    students_left -= 1
                    excel_row += 1
                excel_row = start_row  # Reset to start of the room

            # Move to the next available row after this room
            excel_row = start_row + max(row_benches) + 1
            room_no += 1

            # Print information about remaining students
            if students_left > 0:
                print(f"Students left to be placed: {students_left} (from roll number {current_roll_no} to {starting_roll_no + total_students - 1})")

        # Save the Excel file
        excel_file_name = input("Enter the name of the Excel file: ")
        wb.save(excel_file_name + ".xlsx")

        print("Seating arrangement generated successfully!")
except:
     print(f"you might have given invalid input(i.e string in int)")
     print(f"there might be errors while file handling please run the code again")
